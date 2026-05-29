from __future__ import annotations

import argparse
import math
import textwrap
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


def parse_args():
    p = argparse.ArgumentParser(
        description="Create an annotation-friendly XLSX from qa_label_sheet.csv."
    )
    p.add_argument(
        "--input",
        default="data/processed/qa_label_sheet.csv",
        help="Input label-sheet CSV path.",
    )
    p.add_argument(
        "--output",
        default="data/processed/qa_label_sheet_readable.xlsx",
        help="Output XLSX path.",
    )
    p.add_argument(
        "--wrap-width",
        type=int,
        default=100,
        help="Approximate characters per line for question/answer wrapping.",
    )
    return p.parse_args()


def wrap_cell_text(value: str, width: int) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    return textwrap.fill(text, width=width, break_long_words=False, break_on_hyphens=False)


def estimate_row_height(question: str, answer: str, width: int) -> float:
    q_lines = max(1, math.ceil(len(question) / width))
    a_lines = max(1, math.ceil(len(answer) / width))
    total_lines = min(25, q_lines + a_lines)
    return max(24.0, 14.0 + total_lines * 12.0)


def main():
    args = parse_args()
    input_path = Path(args.input)
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    df = pd.read_csv(input_path)
    for col in ("question", "answer"):
        if col in df.columns:
            df[col] = df[col].map(lambda x: wrap_cell_text(x, args.wrap_width))

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="labels")

    wb = load_workbook(output_path)
    ws = wb["labels"]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Widths for easier reading.
    width_map = {
        "A": 10,   # q_id
        "B": 22,   # company
        "C": 9,    # ticker
        "D": 20,   # date
        "E": 75,   # question
        "F": 85,   # answer
        "G": 22,   # source_url
        "H": 20,   # question_speaker
        "I": 13,   # question_role
        "J": 24,   # answer_speakers
        "K": 13,   # answer_role
        "L": 12,   # label
        "M": 26,   # annotator_notes
    }
    for col, width in width_map.items():
        ws.column_dimensions[col].width = width

    # Header style.
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="top")

    # Wrap text and top-align data rows.
    q_col_idx = None
    a_col_idx = None
    headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
    if "question" in headers:
        q_col_idx = headers.index("question") + 1
    if "answer" in headers:
        a_col_idx = headers.index("answer") + 1

    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).alignment = Alignment(
                wrap_text=True, vertical="top"
            )

        if q_col_idx and a_col_idx:
            q_val = str(ws.cell(row=row, column=q_col_idx).value or "")
            a_val = str(ws.cell(row=row, column=a_col_idx).value or "")
            ws.row_dimensions[row].height = estimate_row_height(
                q_val, a_val, args.wrap_width
            )

    wb.save(output_path)
    print(f"Wrote readable annotation file: {output_path}")


if __name__ == "__main__":
    main()

